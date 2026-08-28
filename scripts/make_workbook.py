#!/usr/bin/env python3
"""Collect every generated table into one Excel workbook.

Groups of per-ligand and per-comparison CSVs are stacked into single sheets
so the workbook has one sheet per analysis rather than one per file, which
is also how the protocol asks results to accumulate as structures are added.
"""
from __future__ import annotations

import csv
import glob
import os
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from mexb_common import REPO, TABLES

OUT = os.path.join(REPO, "results", "MexB_analysis_results.xlsx")
FONT = "Arial"

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
BODY = Font(name=FONT, size=10)
TITLE = Font(name=FONT, bold=True, size=14)
SUB = Font(name=FONT, size=10)
BOLD = Font(name=FONT, bold=True, size=10)
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")


def read(path):
    with open(path) as fh:
        r = list(csv.reader(fh))
    return (r[0], r[1:]) if r else ([], [])


def stack(pattern, keyname, keyfunc):
    """Stack several CSVs sharing a header into one table."""
    files = sorted(glob.glob(os.path.join(TABLES, pattern)))
    header, rows = None, []
    for f in files:
        h, rs = read(f)
        if not h:
            continue
        if header is None:
            header = ([keyname] + h) if keyname else h
        for r in rs:
            rows.append(([keyfunc(f)] + r) if keyname else r)
    return header, rows


def num(v):
    if v is None or v == "":
        return None
    try:
        f = float(v)
        return int(f) if f.is_integer() and abs(f) < 1e15 and "." not in v \
            else f
    except ValueError:
        return v


def add_sheet(wb, name, header, rows, note=None):
    ws = wb.create_sheet(name[:31])
    r0 = 1
    if note:
        ws.cell(1, 1, note).font = Font(name=FONT, italic=True, size=9)
        ws.cell(1, 1).alignment = Alignment(wrap_text=False)
        r0 = 3
    for j, h in enumerate(header, start=1):
        c = ws.cell(r0, j, h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    for i, row in enumerate(rows, start=r0 + 1):
        for j, v in enumerate(row, start=1):
            c = ws.cell(i, j, num(v))
            c.font = BODY
    # widths
    for j, h in enumerate(header, start=1):
        w = len(str(h))
        for row in rows[:400]:
            if j - 1 < len(row):
                w = max(w, len(str(row[j - 1])))
        ws.column_dimensions[get_column_letter(j)].width = min(
            max(w + 2, 9), 46)
    ws.freeze_panes = ws.cell(r0 + 1, 1)
    if rows:
        ws.auto_filter.ref = (f"A{r0}:"
                              f"{get_column_letter(len(header))}"
                              f"{r0 + len(rows)}")
    ws.sheet_view.showGridLines = False
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)

    sheets = []   # (sheet name, description)

    def simple(name, sheet, note=None, desc=""):
        p = os.path.join(TABLES, f"{name}.csv")
        if not os.path.exists(p):
            return
        h, r = read(p)
        add_sheet(wb, sheet, h, r, note)
        sheets.append((sheet, desc, len(r)))

    # ---- README placeholder (filled at the end, must be first sheet)
    readme = wb.create_sheet("README")

    simple("validation", "validation",
           "Automated check against PROTOCOL.md section 6. "
           "Two tunnel checks fail by design - see README and REPORT.md.",
           "Section 6 validation: 49/51 pass")
    simple("inventory", "inventory",
           "Stage 1. Sequence checked against UniProt P52002 at zero offset.",
           "Stage 1 - chains, ranges, gaps, sequence check")
    simple("ligand_inventory", "ligand_inventory",
           "Stage 1. B-factors are group-refined: one value per molecule.",
           "Stage 1 - ligands and their B-factors")
    simple("states", "states",
           "Stage 2. Reference separations (A): Access 26.5/26.2, "
           "Binding 28.3/29.1, Extrusion 29.8/24.5.",
           "Stage 2 - conformational state per protomer")
    simple("proton_relay", "proton_relay",
           "Stage 2. Minimum side-chain functional-atom distances.",
           "Stage 2 - proton relay distance matrix")
    simple("rmsd_protomer_pairs", "rmsd_protomer_pairs",
           "Stage 3. Two superposition frames: trimmed TM, and porter.",
           "Stage 3 - inter-protomer RMSD")
    simple("rmsd_by_region", "rmsd_by_region",
           "Stage 3. TM2/Ialpha displacement and the TM7-12 swing are the "
           "mechanistically important rows.",
           "Stage 3 - per-region deviation")
    simple("rigid_body", "rigid_body",
           "Stage 3. Residual rigid-body transform per domain after the "
           "global fit.",
           "Stage 3 - rigid-body domain motions")
    simple("cross_structure", "cross_structure",
           "Stage 4. Whole-protomer RMSD is an all-CA best fit. Anything "
           "under ~1.0 A is the same state.",
           "Stage 4 - ampicillin vs DDM, same chain")
    simple("ligand_summary", "ligand_summary",
           "Stage 5. One row per bound ligand.",
           "Stage 5 - ligand environment summary")
    simple("contact_matrix", "contact_matrix",
           "Stage 5. Rows are residues, columns are ligands, cells are the "
           "minimum heavy-atom distance in A. Key paper figure.",
           "Stage 5 - cross-ligand contact matrix")

    h, r = stack("contacts_*.csv", None, None)
    if h:
        add_sheet(wb, "ligand_contacts", h, r,
                  "Stage 5. All per-ligand 4.5 A contact lists stacked.")
        sheets.append(("ligand_contacts", "Stage 5 - contacts at 4.5 A",
                       len(r)))
    h, r = stack("hbonds_*.csv", "ligand",
                 lambda f: re.sub(r"^hbonds_|\.csv$", "",
                                  os.path.basename(f)))
    if h:
        add_sheet(wb, "hbonds", h, r,
                  "Stage 5. Candidate hydrogen bonds: N/O to N/O within "
                  "3.6 A.")
        sheets.append(("hbonds", "Stage 5 - candidate hydrogen bonds",
                       len(r)))
    h, r = stack("clashes_*.csv", "ligand",
                 lambda f: re.sub(r"^clashes_|\.csv$", "",
                                  os.path.basename(f)))
    if h:
        add_sheet(wb, "close_contacts", h, r,
                  "Stage 5. Any heavy-atom pair under 2.6 A. The "
                  "interpretation column separates a short hydrogen bond "
                  "from a genuine steric overlap.")
        sheets.append(("close_contacts",
                       "Stage 5 - heavy-atom pairs under 2.6 A", len(r)))

    simple("cavities", "cavities",
           "Stage 6. pyKVFinder, ligand-guided and unguided. Ligands are "
           "stripped for the unguided pass.",
           "Stage 6 - cavity detection")
    simple("pocket_volumes", "pocket_volumes",
           "Stage 6. GRID-BASED volumes: internally comparable across these "
           "structures only, NOT drop-in replacements for fpocket or CASTp.",
           "Stage 6 - substrate-site volume and occlusion")
    simple("pocket_hydropathy", "pocket_hydropathy",
           "Stage 6. Mean Kyte-Doolittle over the PBP- and DBP-lining sets.",
           "Stage 6 - pocket hydropathy")
    simple("pocket_composition", "pocket_composition",
           "Known issue 4. Atom-composition scoring of the two pockets.",
           "Known issue 4 - pocket hydrophobicity")
    simple("fpocket", "fpocket",
           "Stage 6. fpocket on the ligand-stripped trimer; pockets "
           "overlapping a DBP by >= 3 residues.",
           "Stage 6 - fpocket volumes and druggability")
    simple("tunnels", "tunnels",
           "Stage 7. Run on the trimer. 'protein' strips all ligands; "
           "'withlig' keeps them as obstructions. Bottleneck radii are "
           "PROVISIONAL - see README.",
           "Stage 7 - tunnel bottlenecks and channel calls")
    simple("switch_gate", "switch_gate",
           "Diagnostic for the two failing validation checks: how open the "
           "F615 switch-loop gate is, and what a path forced through it "
           "bottlenecks at.",
           "Stage 7 diagnostic - the F615 gate")

    h, r = stack("per_residue_*.csv", "comparison",
                 lambda f: re.sub(r"^per_residue_|\.csv$", "",
                                  os.path.basename(f)))
    if h:
        add_sheet(wb, "per_residue_deviation", h, r,
                  "Stage 3. Sliding-window per-residue CA deviation for "
                  "every protomer pair and both superposition frames, "
                  "stacked. Same data as the ChimeraX .defattr files.")
        sheets.append(("per_residue_deviation",
                       "Stage 3 - per-residue deviation (all pairs)",
                       len(r)))

    fl = os.path.join(TABLES, "flags_all.txt")
    if os.path.exists(fl):
        rows = [[l.strip()] for l in open(fl) if l.strip()]
        add_sheet(wb, "flags", ["flag raised this run"], rows,
                  "Every flag raised by the pipeline on this run.")
        sheets.append(("flags", "Flags raised this run", len(rows)))

    # ---- README
    ws = readme
    ws.sheet_view.showGridLines = False
    ws["A1"] = "MexB substrate-bound structures - tunnel and pocket analysis"
    ws["A1"].font = TITLE
    lines = [
        "",
        "Two cryo-EM structures of the Pseudomonas aeruginosa MexB efflux "
        "transporter, analysed under PROTOCOL.md.",
        "  - Amp_MexB_20260826      ampicillin (ZZ7) bound in chain E",
        "  - MexB_DDM_3_20260730    three DDM (LMT) detergent molecules in "
        "chain E",
        "",
        "Everything here regenerates with 'bash run.sh'. Nothing is "
        "hand-edited.",
        "Full prose write-up with interpretation: results/REPORT.md",
        "",
        "READ THIS BEFORE USING THE NUMBERS",
        "  1. Tunnel bottleneck radii are PROVISIONAL. Two of the 51 "
        "section 6 validation checks fail, both in the tunnel stage: this "
        "implementation",
        "     gives 2.21 A constricted at N676/N718/L827 for ampicillin "
        "chain E, where the protocol expects 2.01 A at F615. Grid "
        "resolution and",
        "     hydrogen handling were both ruled out. The tunnels.py the "
        "protocol says ships with it was not present, so there is no "
        "original",
        "     implementation to compare against. See the switch_gate sheet "
        "and REPORT.md.",
        "  2. CAVER could not be run (academic build is behind a "
        "registration wall), so no tunnel number has been cross-checked "
        "against a second tool.",
        "  3. Grid-based volumes in pocket_volumes are internally "
        "comparable across these two structures only. They are not "
        "equivalent to fpocket or CASTp volumes.",
        "  4. The PBP residue set is flagged as uncertain in the protocol "
        "itself - it came from AcrB literature, not a MexB-specific "
        "source. Any conclusion",
        "     resting on it (pocket assignment, pocket hydropathy) inherits "
        "that uncertainty.",
        "  5. Ligand B-factors are group-refined - one value per molecule - "
        "so they carry no per-atom information.",
        "  6. Detergent is not substrate. Keep the DDM occupancy results "
        "separate from the protein-geometry results.",
        "  7. All CH1/CH2/CH3 channel assignments are tentative; they rest "
        "on lining composition and exit location, not on a reference "
        "channel definition.",
        "",
    ]
    r = 2
    for t in lines:
        c = ws.cell(r, 1, t)
        c.font = BOLD if t == "READ THIS BEFORE USING THE NUMBERS" else SUB
        r += 1
    ws.cell(r, 1, "SHEETS").font = BOLD
    r += 1
    for j, h in enumerate(("sheet", "contents", "rows"), start=1):
        c = ws.cell(r, j, h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
    hdr_row = r
    r += 1
    for name, desc, n in sheets:
        ws.cell(r, 1, name).font = BODY
        ws.cell(r, 1).hyperlink = f"#'{name}'!A1"
        ws.cell(r, 1).font = Font(name=FONT, size=10, color="0563C1",
                                  underline="single")
        ws.cell(r, 2, desc).font = BODY
        ws.cell(r, 3, n).font = BODY
        r += 1
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 62
    ws.column_dimensions["C"].width = 9
    ws.freeze_panes = ws.cell(hdr_row + 1, 1)

    # colour the validation status column
    vs = wb["validation"]
    for row in vs.iter_rows(min_row=4):
        for c in row:
            if c.value == "FAIL":
                c.fill = FAIL_FILL
            elif c.value == "PASS":
                c.fill = PASS_FILL

    wb.save(OUT)
    print(f"wrote {OUT} ({len(sheets) + 1} sheets)")


if __name__ == "__main__":
    main()
